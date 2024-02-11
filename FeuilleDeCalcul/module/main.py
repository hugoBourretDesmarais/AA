# main.py

import os
import re
from datetime import datetime
from collections import defaultdict
import pdfplumber
import warnings
from openpyxl import load_workbook
from openpyxl import Workbook


# Ignore specific UserWarning from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

def main_menu():
    print('''    _    _                 _       _                      _                  _   
   / \  | |_   _ _ __ ___ (_)_ __ (_)_   _ _ __ ___      / \   ___  ___ ___ | |_ 
  / _ \ | | | | | '_ ` _ \| | '_ \| | | | | '_ ` _ \    / _ \ / __|/ __/ _ \| __|
 / ___ \| | |_| | | | | | | | | | | | |_| | | | | | |  / ___ \\__ \ (_| (_) | |_ 
/_/   \_\_|\__,_|_| |_| |_|_|_| |_|_|\__,_|_| |_| |_| /_/   \_\___/\___\___/ \__|
                                                                                 
''')
    while True:
        print("\nMenu Principal")
        print("1. Analyse des fichiers Excel")
        print("0. Quitter")
        
        choice = input("Entrez votre choix: ")
        
        if choice == "1":
            process_excel_files()
        elif choice == "3":
            process_pdf_files()
        elif choice == "0":
            print("Fermeture du programme.")
            break
        else:
            print("Choix invalide, veuillez réessayer.")

def print_results(year, excel_count, mobilisation_sum, outils_sum):
    border = "+" + "-" * 60 + "+"
    title = f" RÉSULTATS DE L'ANALYSE {year} "
    formatted_mobilisation_sum = "${:,.2f}".format(mobilisation_sum)  # With comma as thousand separator, 2 decimal places, and a dollar sign
    formatted_outils_sum = "${:,.2f}".format(outils_sum)

    print(border)
    print(f"|{title.center(60)}|")
    print(border)
    print(f"| Année:                 {str(year).rjust(20)}{' ' * 20}|")
    print(f"| Nombre Total de Projet: {str(excel_count).ljust(20)}{' ' * 20}|")
    print(f"| Mobilisation Total:     {formatted_mobilisation_sum.rjust(20)}{' ' * 20}|")
    print(f"| Outils Total:           {formatted_outils_sum.rjust(20)}{' ' * 20}|")
    print(border)


def find_nearest_facturation_client_dir(file_path):
    """
    Walks up the folder hierarchy from the directory of the given file
    to find the nearest /Facturation - Client/ directory.
    """
    current_dir = os.path.dirname(file_path)
    while current_dir != os.path.dirname(current_dir):  # Check until the root directory
        for item in os.listdir(current_dir):
            if item == "Facturation - Client" and os.path.isdir(os.path.join(current_dir, item)):
                return os.path.join(current_dir, item)
        current_dir = os.path.dirname(current_dir)
    return None

def find_earliest_invoice_date(directory):
    """
    Finds the date of the earliest invoice in the specified directory.
    """
    earliest_date = None
    date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'  # Regex for date format DD/MM/YYYY

    for filename in os.listdir(directory):
        if filename.lower().endswith('.pdf'):
            path = os.path.join(directory, filename)
            try:
                with pdfplumber.open(path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            matches = re.findall(date_pattern, text)
                            for match in matches:
                                date = datetime.strptime(match, "%d/%m/%Y")
                                if earliest_date is None or date < earliest_date:
                                    earliest_date = date
            except Exception as e:
                print(f"Erreur lors du traitement du fichier PDF {filename} : {e}")

    return earliest_date.strftime("%d/%m/%Y") if earliest_date else None


def process_excel_files():
    # Ask the user for the directory path
    directory = input("Veuillez saisir le chemin du répertoire: ")
    
    # Check if the directory exists
    if not os.path.isdir(directory):
        print("Le répertoire fourni n'existe pas. Veuillez entrer un répertoire valide.")
        return

    # A list to hold all aggregated data
    aggregated_data = []

    # Traverse the directory to find Excel files with the pattern '_execute.xlsm'
    for root, dirs, files in os.walk(directory):
        excel_files = [f for f in files if f.lower().endswith('_execute.xlsm')]

        # Process each Excel file
        for filename in excel_files:
            path = os.path.join(root, filename)
            try:
                workbook = load_workbook(filename=path, data_only=True)
                sheet = workbook['Feuille Calcul']
                
                # Use the named range directly. Example: sheet.defined_names['ADMPro'].value
                # You might need a function to get the value from a named range, like this:
                def get_named_range_value(workbook, name):
                    if name in workbook.defined_names:  # Check if the named range exists
                        dest = workbook.defined_names[name].destinations  # Get cell destinations
                        for title, coord in dest:
                            return workbook[title][coord].value
                    return None  # Return None if the named range does not exist
                
                # Access data via named ranges
                data = {
                    'Path': path,
                    'ID': get_named_range_value(workbook, 'SomeNamedRangeForID'),  # Update named range
                    'Nom Projet': get_named_range_value(workbook, 'SomeNamedRangeForNomProjet'),
                    'Date Soumission': get_named_range_value(workbook, 'DateSoumission'),
                    'Date Facturation': get_named_range_value(workbook, 'DateFacturation'),
                    'Pliage': get_named_range_value(workbook, 'Pliage'),
                    'Scellant': get_named_range_value(workbook, 'Scellant'),
                    'Frais Admin': get_named_range_value(workbook, 'ADMIN'),
                    'Outils': get_named_range_value(workbook, 'Outils'),
                    'Mobilisation': get_named_range_value(workbook, 'Mobilisation'),
                    'Frais Dép + Camion': get_named_range_value(workbook, 'FraisDeplacement'),
                    'Remorquage': get_named_range_value(workbook, 'Remorquage'),
                    'Machinerie': get_named_range_value(workbook, 'Machinerie'),
                    'C.P': get_named_range_value(workbook, 'C.P'),
                    'ADM/Pro': get_named_range_value(workbook, 'ADMPro'),
                    'Jours': get_named_range_value(workbook, 'Jours'),
                    'Heures': get_named_range_value(workbook, 'Heures'),
                    'Jour Homme': get_named_range_value(workbook, 'JourHomme'),
                    'Total Installation': get_named_range_value(workbook, 'TotalInstallation'),
                    'Grand Total': get_named_range_value(workbook, 'GrandTotal'),
                    'LivraisonFournisseur': get_named_range_value(workbook, 'LivraisonFournisseur'),
                    'Taux': get_named_range_value(workbook, 'Taux'),
                    'BSDQ': get_named_range_value(workbook, 'BSDQ'),
                    'Entrepreneur': get_named_range_value(workbook, 'Entrepreneur'),
                }
                
                # Add the data to the aggregated list
                aggregated_data.append(data)

            except Exception as e:
                print(f"Erreur lors du traitement du fichier {filename} : {e}")

    # Create a new workbook and select the active sheet
    new_workbook = Workbook()
    dest_sheet = new_workbook.active

    # Write the header to the new sheet
    headers = [
        'Path', 'ID', 'Nom Projet', 'Date Soumission', 'Date Facturation', 'Pliage', 'Scellant',
        'Frais Admin', 'Outils', 'Mobilisation', 'Frais Dép + Camion', 'Remorquage',
        'Machinerie', 'C.P', 'ADM/Pro', 'Jours', 'Heures', 'Jour Homme',
        'Total Installation', 'Grand Total', 'LivraisonFournisseur', 'Taux', 'BSDQ', 'Entrepreneur'
        # Add more headers as needed
    ]
    dest_sheet.append(headers)

    # Write the aggregated data to the new sheet
    for data in aggregated_data:
        row = [data[header] for header in headers]
        dest_sheet.append(row)

    # Save the new workbook
    new_workbook.save(os.path.join(directory, 'aggregated_data.xlsx'))
    print("Données agrégées avec succès.")

def process_pdf_files():
    directory = input("Veuillez saisir le chemin du répertoire pour les fichiers PDF: ")
    if not os.path.isdir(directory):
        print("Le répertoire fourni n'existe pas. Veuillez entrer un répertoire valide.")
        return

    date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'  # Regex for date format DD/MM/YYYY
    found_dates = []

    # Use os.walk to recursively traverse the directories
    for root, dirs, files in os.walk(directory):
        # Filter files to include only PDF files
        pdf_files = [f for f in files if f.lower().endswith('.pdf')]

        for filename in pdf_files:
            path = os.path.join(root, filename)
            try:
                with pdfplumber.open(path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            matches = re.findall(date_pattern, text)
                            found_dates.extend(matches)
            except Exception as e:
                print(f"Erreur lors du traitement du fichier PDF {filename} : {e}")

    print("Dates trouvées dans les fichiers PDF:")
    for date in set(found_dates):
        print(date)



if __name__ == "__main__":
    main_menu()
1